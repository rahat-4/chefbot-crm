import calendar
import datetime
import io
import openpyxl
import phonenumbers


def generate_excel(title: str, headers: list, data: list) -> io.BytesIO:
    """
    Generate an Excel file from the provided data and headers.

    :param data: List of dictionaries containing the data to be written to the Excel file.
    :param headers: List of strings representing the column headers.
    :return: BytesIO object containing the Excel file data.
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, row_data in enumerate(data, start=2):
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=row_data.get(header))

    # Save the workbook to a BytesIO stream
    excel_stream = io.BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)  # Reset stream position to the beginning

    return excel_stream


def get_timestamped_filename(base_name, extension="xlsx"):
    """
    Generate a timestamped filename.

    :param base_name: Base name for the file.
    :param extension: File extension (default is 'xlsx').
    :return: String representing the timestamped filename.
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{timestamp}.{extension}"


def format_phone_number(phone: str) -> str:
    if phone and isinstance(phone, phonenumbers.PhoneNumber):
        return phonenumbers.format_number(
            phone, phonenumbers.PhoneNumberFormat.INTERNATIONAL
        )
    return phone or ""


def ordinal(n):
    # Returns the ordinal suffix for a given day number
    return f"{n}{'th' if 11<=n<=13 else {1:'st', 2:'nd', 3:'rd'}.get(n%10, 'th')}"


def format_day_month(d):
    if not d:
        return ""
    # Get day and month
    day = d.day
    month = calendar.month_name[d.month]
    return f"{ordinal(day)} {month}"


# Example usage:
if __name__ == "__main__":
    sample_data = [
        {"Name": "Alice", "Age": 30, "City": "New York"},
        {"Name": "Bob", "Age": 25, "City": "Los Angeles"},
        {"Name": "Charlie", "Age": 35, "City": "Chicago"},
    ]
    headers = ["Name", "Age", "City"]

    excel_file = generate_excel(sample_data, headers)
    filename = get_timestamped_filename("user_data")

    # Save to a file for demonstration purposes
    with open(filename, "wb") as f:
        f.write(excel_file.getbuffer())

    print(f"Excel file '{filename}' generated successfully.")
